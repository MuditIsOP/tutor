from datetime import date
import csv
from pathlib import Path

from sqlalchemy import text
from sqlalchemy.orm import Session

from models import Course, Module, Subject, Topic


COURSE_SEED_DATA = [
    {"course_id": 1, "course_name": "B.Tech", "total_years": 4},
    {"course_id": 2, "course_name": "BBA", "total_years": 3},
    {"course_id": 3, "course_name": "BSc", "total_years": 3},
]
BOOK1_PATH = Path(__file__).resolve().parent.parent / "Book1.xlsx"
MODULES_CSV_PATH = Path(__file__).resolve().parent.parent / "modules_cleaned.csv"


def seed_courses(db: Session) -> None:
    for course_data in COURSE_SEED_DATA:
        existing_course = db.get(Course, course_data["course_id"])
        if existing_course:
            continue

        db.add(Course(**course_data))

    db.commit()


def migrate_students_table(db: Session) -> None:
    columns = {
        row[1] for row in db.execute(text("PRAGMA table_info(students)")).fetchall()
    }
    if not columns:
        return

    if "branch_code" in columns:
        db.execute(text("ALTER TABLE students RENAME TO students_legacy"))
        db.execute(
            text(
                """
                CREATE TABLE students (
                    student_id VARCHAR(32) PRIMARY KEY,
                    name VARCHAR(120) NOT NULL,
                    dob DATE NOT NULL,
                    gender VARCHAR(20) NOT NULL,
                    email VARCHAR(255) NOT NULL UNIQUE,
                    phone VARCHAR(20) NOT NULL,
                    course_id INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    password_hash VARCHAR(64) NOT NULL,
                    profile_photo VARCHAR(255),
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(course_id) REFERENCES courses (course_id)
                )
                """
            )
        )
        db.execute(
            text(
                """
                INSERT INTO students (
                    student_id, name, dob, gender, email, phone, course_id,
                    year, password_hash, profile_photo, created_at, updated_at
                )
                SELECT
                    student_id, name, dob, gender, email, phone, course_id,
                    year, password_hash, profile_photo, created_at, updated_at
                FROM students_legacy
                """
            )
        )
        db.execute(text("DROP TABLE students_legacy"))
        db.commit()
        columns = {
            row[1] for row in db.execute(text("PRAGMA table_info(students)")).fetchall()
        }

    if "semester" not in columns:
        db.execute(text("ALTER TABLE students RENAME TO students_pre_semester"))
        db.execute(
            text(
                """
                CREATE TABLE students (
                    student_id VARCHAR(32) PRIMARY KEY,
                    name VARCHAR(120) NOT NULL,
                    dob DATE NOT NULL,
                    gender VARCHAR(20) NOT NULL,
                    email VARCHAR(255) NOT NULL UNIQUE,
                    phone VARCHAR(20) NOT NULL,
                    course_id INTEGER NOT NULL,
                    year INTEGER NOT NULL,
                    semester INTEGER NOT NULL,
                    password_hash VARCHAR(64) NOT NULL,
                    profile_photo VARCHAR(255),
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY(course_id) REFERENCES courses (course_id)
                )
                """
            )
        )
        db.execute(
            text(
                """
                INSERT INTO students (
                    student_id, name, dob, gender, email, phone, course_id,
                    year, semester, password_hash, profile_photo, created_at, updated_at
                )
                SELECT
                    student_id, name, dob, gender, email, phone, course_id,
                    year, year * 2, password_hash, profile_photo, created_at, updated_at
                FROM students_pre_semester
                """
            )
        )
        db.execute(text("DROP TABLE students_pre_semester"))
        db.commit()

    students = db.execute(
        text(
            """
            SELECT student_id, course_id, created_at
            FROM students
            ORDER BY course_id ASC, datetime(created_at) ASC, student_id ASC
            """
        )
    ).fetchall()

    counters: dict[int, int] = {}
    for student_id, course_id, created_at in students:
        counters[course_id] = counters.get(course_id, 0) + 1
        registration_year = date.fromisoformat(str(created_at)[:10]).year
        normalized_id = f"STD{registration_year}{course_id:02d}{counters[course_id]:03d}"
        if normalized_id != student_id:
            db.execute(
                text("UPDATE students SET student_id = :new_id WHERE student_id = :old_id"),
                {"new_id": normalized_id, "old_id": student_id},
            )

    db.commit()


def import_subjects_from_excel(db: Session) -> None:
    if not BOOK1_PATH.exists():
        return

    from openpyxl import load_workbook

    workbook = load_workbook(BOOK1_PATH, data_only=True)
    worksheet = workbook.active
    headers = [str(cell).strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue

        record = dict(zip(headers, row))
        subject = db.get(Subject, str(record["subject_code"]).strip())
        payload = {
            "subject_code": str(record["subject_code"]).strip(),
            "subject_name": str(record["subject_name"]).strip(),
            "credits": int(record["credits"]),
            "type": str(record["type"]).strip(),
            "semester": int(record["semester"]),
            "year": int(record["year"]),
            "course_id": int(record["course_id"]),
        }

        if subject:
            for key, value in payload.items():
                setattr(subject, key, value)
            continue

        db.add(Subject(**payload))

    db.commit()


def import_modules_from_csv(db: Session) -> None:
    if not MODULES_CSV_PATH.exists():
        return

    with MODULES_CSV_PATH.open(newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.DictReader(csv_file)
        for row in reader:
            if not row.get("subject_code") or not row.get("module_number") or not row.get("module_title"):
                continue

            subject_code = str(row["subject_code"]).strip()
            module_number = int(row["module_number"])
            module_title = str(row["module_title"]).strip()
            module_id = f"{subject_code}_{module_number}"

            module = db.get(Module, module_id)
            payload = {
                "module_id": module_id,
                "subject_code": subject_code,
                "module_number": module_number,
                "module_title": module_title,
            }

            if module:
                for key, value in payload.items():
                    setattr(module, key, value)
                continue

            db.add(Module(**payload))

    db.commit()


def import_topics_from_csv(db: Session) -> None:
    if not MODULES_CSV_PATH.exists():
        return

    columns = {
        row[1] for row in db.execute(text("PRAGMA table_info(topics)")).fetchall()
    }
    if not columns:
        db.commit()
    elif "topic_id" not in columns:
        db.execute(text("DROP TABLE topics"))
        db.commit()
        return import_topics_from_csv(db)

    db.execute(text("DELETE FROM topics"))

    with MODULES_CSV_PATH.open(newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.DictReader(csv_file)
        for row in reader:
            if not row.get("subject_code") or not row.get("module_number"):
                continue

            subject_code = str(row["subject_code"]).strip()
            module_number = int(row["module_number"])
            module_id = f"{subject_code}_{module_number}"
            raw_topics = str(row.get("topics") or "")
            topic_values = [topic.strip() for topic in raw_topics.split(";") if topic.strip()]

            for index, topic in enumerate(topic_values, start=1):
                db.add(Topic(topic_id=f"{module_id}_{index}", module_id=module_id, topic=topic))

    db.commit()
